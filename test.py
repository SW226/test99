import json
import sys
import time
import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials as SAC 
from openpyxl import load_workbook
from openpyxl import Workbook



GDriveJSON = 'midyear-calling-270711-028f9ad25b9a.json'
GSpreadSheet = '庫存表'
while True:
    try:
        scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
        key = SAC.from_json_keyfile_name(GDriveJSON, scope)
        gc = gspread.authorize(key)
        worksheet = gc.open(GSpreadSheet).sheet1
    except Exception as ex:
        print('無法連線Google試算表', ex)
        line_bot_api.reply_message(
        event.reply_token,
        TextSendMessage(text='無法連線Google試算表'))
        sys.exit(1)

    wb = load_workbook('1.xlsx',data_only=True)

    sheet_names = wb.sheetnames

    ws = wb[sheet_names[0]] 
    ws1 = []
    rows = ws.rows
    columns = ws.columns
    for row in rows:
        lines = []
        line = [col.value for col in row]
        if line[1] != '' and line[1] != '品名':
            lines.append(line[0])
            lines.append(line[1])
            lines.append(line[8])
            lines.append(line[5])
            worksheet.append_row(lines)
    break